import openpyxl, json

wb = openpyxl.load_workbook('keywords.xlsx', data_only=True)
data = {} ; text1 = [] ; text2 = []

for mname in wb.sheetnames:
    ws = wb[mname]
    module = data[mname] = {}
    command = []
    for row in ws.iter_rows(values_only=True):
        for i, key in enumerate(row):
            if key is None: continue
            if isinstance(key, str) and len(key) == 0: continue
            
            if i == 0:
                command = module[key] = []
            else:
                command.append(key)

with open('assets/keywords.json', 'w') as f:
    json.dump(data, f, indent=2)

for mname, module in data.items():
    if mname=='BASIC': continue
    text1.append(rf'''
  {{
  begin: '(?i)^ *([\\$\\+-]?PROG)( +{mname})( .*)?$'
  beginCaptures:
    1: name: 'support.class.{mname.lower()}.sofistik'
    2: name: 'support.class.{mname.lower()}.sofistik'
    3:
      name: 'comment.line.{mname.lower()}.sofistik'
      patterns: [{{ include: 'text.todo' }}]
  end: '(?i)(?=^ *([\\$\\+-]?PROG))'
  name: 'module.{mname.lower()}.sofistik'
  patterns: [
    {{'include': '#{mname}'}}
    ]
  }}''')



    # Flattened structure: One rule for all commands, one rule for all sub-keywords
    all_commands = set()
    all_subkeywords = set()
    
    for cname, subcmds in module.items():
        all_commands.add(cname)
        all_subkeywords.update(subcmds)
    
    # Sort for consistent regex (longest first to avoid prefix matching issues)
    sorted_commands = sorted(list(all_commands), key=lambda x: (-len(x), x))
    sorted_subkeywords = sorted(list(all_subkeywords), key=lambda x: (-len(x), x))
    
    # Create optimized regex strings
    cmds_pattern = "|".join(sorted_commands)
    subcmds_pattern = "|".join(sorted_subkeywords)
    
    text2.append([rf'''
  {mname}: {{
    patterns: [
      {{ include: '#defA' }}
      {{
        match: '(?i)(?:^ *|; *)({cmds_pattern})(?=;|$| )'
        captures:
          1: name: 'keyword.control.sofistik'
      }}
      {{
        match: '(?i)(?<!\\w)({subcmds_pattern})(?!\\w)'
        name: 'entity.name.function.sofistik'
      }}
      {{ include: '#normalText' }}
    ]
  }}'''])

text = (r'''
# ***** References *****
# https://pulsar-edit.dev/docs/launch-manual/sections/core-hacking/#creating-a-legacy-textmate-grammar
# https://gist.github.com/savetheclocktower/c9607b97477d4817911e4f2f8db89679
# http://manual.macromates.com/en/language_grammars/
# https://github.com/kkos/oniguruma/blob/master/doc/RE
# https://regex101.com/

scopeName: 'source.sofistik'
name: 'SOFiSTiK'
fileTypes: ['dat','gra','grb','results']
patterns: [
  {
    match: '(?i)^@ *SOFiSTiK *(\\d{4})(-\\d\\d?)? *$'
    name: 'meta.version.sofistik'
  }
  {
    match: '(?i)^@ .+'
    name: 'meta.sofistik'
  }
  { include: '#normalText' }
'''+

''.join(text1)

+r'''
]
repository:
  normalText: {
    patterns: [
      { include: '#textcmds' }
      { include: '#defA' }
      { include: '#defB' }
      { include: '#defC' }
      { include: '#defD' }
      { include: '#inc1' }
      { include: '#text' }
      { include: '#comments' }
      { include: '#str1' }
      { include: '#str2' }
      { include: '#var1' }
      { include: '#var2' }
      { include: '#keys1' }
      { include: '#keys2' }
      { include: '#expr' }
      { include: '#units' }
    ]
  }
  textcmds: {
    match: '(?i)(^ *)(HEAD|TXB|TXE)( .+?$| *$)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  text: {
    begin: '(?i)^[ ]*(<TEXT>|<TEXT,FILE=\\+?(.+)>)(?= |$)'
    beginCaptures:
      1: name: 'support.function.sofistik'
      2: name: 'string.other.sofistik'
    end: '(?i)^[ ]*(<\\/TEXT>)(?= |$)'
    endCaptures:
      1: name: 'support.function.sofistik'
    patterns: [
      { include: '#inc1' }
      { include: '#var2' }
      { include: '#edit' }
    ]
  }
  edit: {
    begin: '(?i)(<EDIT:.+?>)'
    beginCaptures:
      1: name: 'support.function.sofistik'
    end: '(?i)(<\\/EDIT>)'
    endCaptures:
      1: name: 'support.function.sofistik'
    patterns: [
      { include: '#var2' }
      { include: '#inc1' }
    ]
  }
  defA: {
    match: '(?i)^[ ]*(#DEFINE|#ENDDEF) *(.+?)?(?: *= *(.*))?$'
    captures:
      1: name: 'entity.name.section.sofistik'
      2: name: 'string.other.sofistik'
      3: patterns: [
        { include: '#inc1' }
        { include: '#comments' }
        { include: '#str1' }
        { include: '#str2' }
        { include: '#var1' }
        { include: '#var2' }
        { include: '#keys1' }
        { include: '#keys2' }
        { include: '#expr' }
        { include: '#units' }
      ]
      4: name: 'entity.name.section.sofistik'
  }
  defB: {
    match: '(?i)^[ ]*(#INCLUDE) +(.+)'
    captures:
      1: name: 'entity.name.section.sofistik'
      2:
        name: 'string.other.sofistik'
        patterns: [{ include: '#inc1' }]
  }
  defC: {
    match: '(?i)^[ ]*([\\$\\+-]?APPLY|[\\+-]?SYS)( +.+)'
    captures:
      1: name: 'support.class.sofistik'
      2:
        name: 'string.other.sofistik'
        patterns: [
          { include: '#inc1' }
          { include: '#comments' }
        ]
  }
  defD: {
    match: '(?i)^[ ]*(#IF|#ELSE|#ENDIF)'
    captures:
      1: name: 'entity.name.section.sofistik'
  }
  inc1: {
    match: '(;)?[ ]*(\\$\\(\\S+?\\))'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'variable.other.sofistik'
  }
  comments: {
    match: '(?i)(?:!|\\/\\/|\\$(?!PROG))(.*)'
    name: 'comment.line.sofistik'
    captures:
      1: patterns: [{ include: 'text.todo' }]
  }
  str1: {
    match: '\\"(.*?)\\"'
    name: 'string.double.sofistik'
    captures:
      1: patterns: [{ include: '#inc1' }]
  }
  str2: {
    match: "\\'(.*?)\\'"
    name: 'string.single.sofistik'
    captures:
      1: patterns: [{ include: '#inc1' }]
  }
  var1: {
    match: '(?i)(^|;)[ ]*(LET|STO|DEL|DBG|PRT)(?!\\w)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  var2: {
    match: '(#\\w+|#\\(\\w+(?:,\\d\\.\\d)?\\))'
    name: 'variable.other.sofistik'
  }
  keys1: {
    match: '(?i)(^|;)[ ]*(LOOP)(?!\\w)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  keys2: {
    match: '(?i)(^|;)[ ]*(IF|ELSEIF|ELSE|ENDIF|ENDLOOP|END)(?=\\s|$)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  expr: {
    match: '(?<=\\s|^)(=\\S+)'
    captures:
      1:
        name: 'entity.name.function.sofistik'
        patterns: [{'include': '#var2'}]
  }
  units: {
    match: '(?<=\\S)\\[.*?\\]'
    name: 'constant.other.sofistik'
  }
''' +

''.join([''.join(txt) for txt in text2])
)

with open('grammars/sofistik.cson', 'w') as f:
    f.write(text[1:]+'\n')
