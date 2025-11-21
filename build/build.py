"""
SOFiSTiK Grammar Builder - Fixed Version v4
Now properly scopes subkeywords to their parent commands
Scope continues across ALL lines until another command appears or semicolon
FIXED: Module scope (PROG) now works correctly
"""

import openpyxl
import json
from collections import defaultdict


def load_keywords(excel_path='keywords.xlsx'):
    """Load keywords from Excel file and organize by module."""
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    keywords_by_module = {}

    for module_name in workbook.sheetnames:
        worksheet = workbook[module_name]
        module_keywords = keywords_by_module[module_name] = {}
        current_command_subkeywords = []

        for row in worksheet.iter_rows(values_only=True):
            for column_index, cell_value in enumerate(row):
                if cell_value is None or (isinstance(cell_value, str) and len(cell_value) == 0):
                    continue

                if column_index == 0:
                    # First column contains command names
                    current_command_subkeywords = module_keywords[cell_value] = []
                else:
                    # Other columns contain sub-keywords for the command
                    current_command_subkeywords.append(cell_value)

    return keywords_by_module


def save_keywords_json(keywords_data, output_path='../assets/keywords.json'):
    """Save keywords as JSON for programmatic access."""
    with open(output_path, 'w') as json_file:
        json.dump(keywords_data, json_file, indent=2)


def optimize_pattern(keyword_list):
    """
    Create optimized regex pattern by grouping keywords by first character.
    This reduces backtracking and improves matching performance.

    Args:
        keyword_list: List of keywords to optimize into a regex pattern

    Returns:
        Optimized regex pattern string
    """
    if not keyword_list:
        return ""

    # Group keywords by their first character (case-insensitive)
    keywords_grouped_by_first_char = defaultdict(list)
    for keyword in keyword_list:
        first_character = keyword[0].upper()
        keywords_grouped_by_first_char[first_character].append(keyword)

    # Sort within each group by length (longest first) to avoid prefix matching issues
    for character in keywords_grouped_by_first_char:
        keywords_grouped_by_first_char[character].sort(key=lambda x: (-len(x), x))

    # Build optimized pattern
    if len(keywords_grouped_by_first_char) == 1:
        # Single character group - use simple alternation
        character, word_list = list(keywords_grouped_by_first_char.items())[0]
        return "|".join(word_list)

    # Multiple character groups - use character class optimization
    pattern_parts = []
    for character in sorted(keywords_grouped_by_first_char.keys()):
        word_list = keywords_grouped_by_first_char[character]
        if len(word_list) == 1:
            pattern_parts.append(word_list[0])
        else:
            # Group words with same first character
            word_suffixes = '|'.join(word[1:] for word in word_list)
            pattern_parts.append(f"(?:{character}(?:{word_suffixes}))")

    return "|".join(pattern_parts)


def generate_module_patterns(keywords_data):
    """
    Generate optimized pattern blocks for each SOFiSTiK module.
    NOW PROPERLY SCOPES SUBKEYWORDS TO THEIR PARENT COMMANDS.
    Scope continues across ALL lines until another command or semicolon.

    Args:
        keywords_data: Dictionary of modules and their keywords

    Returns:
        Tuple of (module_pattern_list, repository_pattern_list)
    """
    module_begin_end_patterns = []
    module_repository_patterns = []

    for module_name, module_commands in keywords_data.items():
        if module_name == 'BASIC':
            continue

        module_name_lowercase = module_name.lower()

        # Generate module begin/end pattern for PROG statements
        module_begin_end_patterns.append(f'''
  {{
  begin: '(?i)^ *([\\\\$\\\\+-]?PROG)( +{module_name})( .*)?$'
  beginCaptures:
    1: name: 'support.class.{module_name_lowercase}.sofistik'
    2: name: 'support.class.{module_name_lowercase}.sofistik'
    3:
      name: 'comment.line.{module_name_lowercase}.sofistik'
      patterns: [{{ include: 'text.todo' }}]
  end: '(?i)(?=^ *[\\\\$\\\\+-]?PROG\\\\b)'
  name: 'module.{module_name_lowercase}.sofistik'
  patterns: [
    {{'include': '#{module_name}'}}
    ]
  }}''')

        # Collect all commands for this module
        all_module_commands = list(module_commands.keys())

        # Sort by length (longest first) to prevent prefix matching issues
        sorted_commands = sorted(all_module_commands, key=lambda x: (-len(x), x))

        # Generate optimized regex pattern for all commands - to use in the end pattern
        commands_regex_pattern = optimize_pattern(sorted_commands)

        # Build repository entry with command-specific subkeyword scoping
        repository_entry = f'''
  {module_name}: {{
    patterns: [
      {{ include: '#preprocessorDefine' }}'''

        # Generate a begin/end pattern for EACH command with its specific subkeywords
        for command_name, command_subkeywords in module_commands.items():
            if command_subkeywords:  # Only if command has subkeywords
                # Sort subkeywords by length (longest first)
                sorted_subkeywords = sorted(command_subkeywords, key=lambda x: (-len(x), x))
                subkeywords_regex_pattern = optimize_pattern(sorted_subkeywords)

                # Create a scoped pattern for this command and its subkeywords
                # The scope continues across ALL lines until:
                # 1. Another command from this module appears (at line start or after semicolon)
                # 2. A semicolon is encountered (ends current command)
                # 3. A new PROG statement (which ends the module scope)
                repository_entry += f'''
      {{
        begin: '(?i)(?:^ *|; *)({command_name})(?=;|$| )'
        beginCaptures:
          1: name: 'keyword.control.sofistik'
        end: '(?i)(?=(?:^ *|; *)(?:{commands_regex_pattern})|;|^ *[\\\\$\\\\+-]?PROG\\\\b)'
        patterns: [
          {{
            match: '(?i)(?<!\\\\w)({subkeywords_regex_pattern})(?!\\\\w)'
            name: 'entity.name.function.sofistik'
          }}
          {{ include: '#normalText' }}
        ]
      }}'''
            else:
                # Command without subkeywords - just match it directly
                repository_entry += f'''
      {{
        match: '(?i)(?:^ *|; *)({command_name})(?=;|$| )'
        captures:
          1: name: 'keyword.control.sofistik'
      }}'''

        # Close the repository entry
        repository_entry += '''
      { include: '#normalText' }
    ]
  }'''

        module_repository_patterns.append(repository_entry)

    return module_begin_end_patterns, module_repository_patterns


def generate_grammar(keywords_data):
    """
    Generate complete optimized CSON grammar file.

    Args:
        keywords_data: Dictionary of modules and their keywords

    Returns:
        Complete grammar as string
    """
    module_patterns, repository_patterns = generate_module_patterns(keywords_data)

    # Grammar header with metadata and references
    grammar_header = r'''
# ***** References *****
# https://pulsar-edit.dev/docs/launch-manual/sections/core-hacking/#creating-a-legacy-textmate-grammar
# https://gist.github.com/savetheclocktower/c9607b97477d4817911e4f2f8db89679
# http://manual.macromates.com/en/language_grammars/
# https://github.com/kkos/oniguruma/blob/master/doc/RE
# https://regex101.com/

scopeName: 'source.sofistik'
name: 'SOFiSTiK'
fileTypes: ['dat','gra','grb','results']
patterns: [
  {
    match: '(?i)^@ *SOFiSTiK *(\\d{4})(-\\d\\d?)? *$'
    name: 'meta.version.sofistik'
  }
  {
    match: '(?i)^@ .+'
    name: 'meta.sofistik'
  }
  { include: '#normalText' }
'''

    # Add module-specific patterns
    grammar_with_modules = grammar_header + ''.join(module_patterns)

    # Repository with base patterns for common SOFiSTiK syntax
    repository_base_patterns = r'''
]
repository:
  normalText: {
    patterns: [
      { include: '#textCommands' }
      { include: '#preprocessorDefine' }
      { include: '#preprocessorInclude' }
      { include: '#preprocessorApplyAndSys' }
      { include: '#preprocessorConditional' }
      { include: '#dollarVariableReference' }
      { include: '#textBlock' }
      { include: '#comments' }
      { include: '#stringDoubleQuoted' }
      { include: '#stringSingleQuoted' }
      { include: '#variableCommands' }
      { include: '#hashVariableReference' }
      { include: '#loopKeyword' }
      { include: '#controlFlowKeywords' }
      { include: '#expressionWithEquals' }
      { include: '#units' }
    ]
  }
  textCommands: {
    match: '(?i)(^ *)(HEAD|TXB|TXE)( .+?$| *$)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  textBlock: {
    begin: '(?i)^[ ]*(<TEXT>|<TEXT,FILE=\\+?(.+)>)(?= |$)'
    beginCaptures:
      1: name: 'support.function.sofistik'
      2: name: 'string.other.sofistik'
    end: '(?i)^[ ]*(<\\/TEXT>)(?= |$)'
    endCaptures:
      1: name: 'support.function.sofistik'
    patterns: [
      { include: '#dollarVariableReference' }
      { include: '#hashVariableReference' }
      { include: '#editBlock' }
    ]
  }
  editBlock: {
    begin: '(?i)(<EDIT:.+?>)'
    beginCaptures:
      1: name: 'support.function.sofistik'
    end: '(?i)(<\\/EDIT>)'
    endCaptures:
      1: name: 'support.function.sofistik'
    patterns: [
      { include: '#hashVariableReference' }
      { include: '#dollarVariableReference' }
    ]
  }
  preprocessorDefine: {
    match: '(?i)^[ ]*(#DEFINE|#ENDDEF) *(.+?)?(?: *= *(.*))?$'
    captures:
      1: name: 'entity.name.section.sofistik'
      2: name: 'string.other.sofistik'
      3: patterns: [
        { include: '#dollarVariableReference' }
        { include: '#comments' }
        { include: '#stringDoubleQuoted' }
        { include: '#stringSingleQuoted' }
        { include: '#variableCommands' }
        { include: '#hashVariableReference' }
        { include: '#loopKeyword' }
        { include: '#controlFlowKeywords' }
        { include: '#expressionWithEquals' }
        { include: '#units' }
      ]
      4: name: 'entity.name.section.sofistik'
  }
  preprocessorInclude: {
    match: '(?i)^[ ]*(#INCLUDE) +(.+)'
    captures:
      1: name: 'entity.name.section.sofistik'
      2:
        name: 'string.other.sofistik'
        patterns: [{ include: '#dollarVariableReference' }]
  }
  preprocessorApplyAndSys: {
    match: '(?i)^[ ]*([\\$\\+-]?APPLY|[\\+-]?SYS)( +.+)'
    captures:
      1: name: 'support.class.sofistik'
      2:
        name: 'string.other.sofistik'
        patterns: [
          { include: '#dollarVariableReference' }
          { include: '#comments' }
        ]
  }
  preprocessorConditional: {
    match: '(?i)^[ ]*(#IF|#ELSE|#ENDIF)'
    captures:
      1: name: 'entity.name.section.sofistik'
  }
  dollarVariableReference: {
    match: '(;)?[ ]*(\\$\\(\\S+?\\))'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'variable.other.sofistik'
  }
  comments: {
    match: '(?i)(?:!|\\/\\/|\\$(?!PROG))(.*)'
    name: 'comment.line.sofistik'
    captures:
      1: patterns: [{ include: 'text.todo' }]
  }
  stringDoubleQuoted: {
    match: '\\"(.*?)\\"'
    name: 'string.double.sofistik'
    captures:
      1: patterns: [{ include: '#dollarVariableReference' }]
  }
  stringSingleQuoted: {
    match: "\\'(.*?)\\'"
    name: 'string.single.sofistik'
    captures:
      1: patterns: [{ include: '#dollarVariableReference' }]
  }
  variableCommands: {
    match: '(?i)(^|;)[ ]*(LET|STO|DEL|DBG|PRT)(?!\\w)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  hashVariableReference: {
    match: '(#\\w+|#\\(\\w+(?:,\\d\\.\\d)?\\))'
    name: 'variable.other.sofistik'
  }
  loopKeyword: {
    match: '(?i)(^|;)[ ]*(LOOP)(?!\\w)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  controlFlowKeywords: {
    match: '(?i)(^|;)[ ]*(IF|ELSEIF|ELSE|ENDIF|ENDLOOP|END)(?=\\s|$)'
    captures:
      1: name: 'support.type.sofistik'
      2: name: 'keyword.control.sofistik'
  }
  expressionWithEquals: {
    match: '(?<=\\s|^)(=\\S+)'
    captures:
      1:
        name: 'entity.name.function.sofistik'
        patterns: [{'include': '#hashVariableReference'}]
  }
  units: {
    match: '(?<=\\S)\\[.*?\\]'
    name: 'constant.other.sofistik'
  }
'''

    # Combine all parts: header + modules + repository + module repositories
    complete_grammar = grammar_with_modules + repository_base_patterns + ''.join(repository_patterns)

    return complete_grammar


def main():
    """Main build process - orchestrates the grammar generation."""
    print("Loading keywords from Excel...")
    keywords_data = load_keywords()

    print("Saving keywords.json...")
    save_keywords_json(keywords_data)

    print("Generating optimized grammar with properly scoped subkeywords...")
    complete_grammar = generate_grammar(keywords_data)

    print("Writing grammar file...")
    with open('../grammars/sofistik.cson', 'w', encoding='utf-8') as grammar_file:
        grammar_file.write(complete_grammar[1:] + '\n')  # Remove leading newline

    print("Build complete!")
    print(f"- Generated grammar for {len(keywords_data)} modules")
    total_command_count = sum(len(module) for module in keywords_data.values())
    print(f"- Total commands: {total_command_count}")


if __name__ == '__main__':
    main()
